"""
json_to_excel.py — تحويل JSON إلى Excel منظم وملون
====================================================
يحول ملف products_full.json إلى ملف Excel احترافي مع:
  ✨ رؤوس ملونة جذابة
  ✨ ترتيب وتنسيق احترافي
  ✨ ألوان متبادلة للصفوف
  ✨ أعمدة محسّنة (الصور، الروابط)
  ✨ ترتيب تلقائي حسب السعر والمبيعات

التثبيت:
  pip install openpyxl pandas

التشغيل:
  python json_to_excel.py
"""

import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink

# ============================================================
# CONFIGURATION
# ============================================================

INPUT_FILE  = "results/products_full.json"
OUTPUT_FILE = "results/products_catalog.xlsx"

# ألوان جذابة
HEADER_COLOR    = "1F4E78"      # أزرق غامق احترافي
HEADER_TEXT     = "FFFFFF"      # أبيض
STRIPE_COLOR_1  = "FFFFFF"      # أبيض
STRIPE_COLOR_2  = "E8F0F7"      # أزرق فاتح جداً
ACCENT_COLOR    = "4472C4"      # أزرق متوسط للبيانات المهمة

# ============================================================
# HELPER FUNCTIONS
# ============================================================

def calculate_savings(product):
    """حساب نسبة الخصم والتوفير"""
    try:
        regular = float(str(product.get("regular_price", "0")).replace(",", "."))
        special = float(str(product.get("special_price", product.get("price", "0"))).replace(",", "."))
        
        if regular > special > 0:
            discount = ((regular - special) / regular) * 100
            return f"-{discount:.0f}% ({regular - special:.2f} ر.س)"
        return ""
    except:
        return ""

# ============================================================
# LOAD & PROCESS DATA
# ============================================================

print("[INFO] جاري تحميل البيانات...")
with open(INPUT_FILE, "r", encoding="utf-8") as f:
    data = json.load(f)

print(f"[INFO] وُجد {len(data)} منتج")

# تحضير البيانات للـ DataFrame
records = []
for product in data:
    # استخراج أول صورة
    first_image = product.get("images", [None])[0] if product.get("images") else None
    images_count = len(product.get("images", []))
    
    # تحويل السعر والمبيعات إلى أرقام
    regular_price_str = str(product.get("regular_price", ""))
    special_price_str = str(product.get("special_price", product.get("price", "0")))
    
    # تنظيف السعر الأصلي
    try:
        regular_price = float(regular_price_str.replace(",", ".").split()[0]) if regular_price_str else 0.0
    except:
        regular_price = 0.0
    
    # تنظيف السعر النهائي
    try:
        special_price = float(special_price_str.replace(",", ".").split()[0]) if special_price_str else 0.0
    except:
        special_price = 0.0
    
    # تحديد السعر المستخدم للترتيب
    price = special_price if special_price > 0 else regular_price
    
    sales_str = str(product.get("sales", "0"))
    try:
        sales = int(sales_str.replace(" ", ""))
    except:
        sales = 0
    
    # اختصار الوصف (أول 150 حرف)
    description = product.get("description", "")
    short_desc = description[:150] + "..." if len(description) > 150 else description
    
    records.append({
        "ID": product.get("id"),
        "اسم المنتج": product.get("name", ""),
        "السعر الأصلي": product.get("regular_price", ""),
        "السعر النهائي": product.get("special_price") or product.get("price", ""),
        "التوفير": calculate_savings(product),
        "SKU": product.get("sku", ""),
        "المبيعات": sales,
        "عدد الصور": images_count,
        "الصورة الأولى": first_image or "لا توجد",
        "الوصف المختصر": short_desc,
        "الرابط": product.get("url", ""),
    })

# إنشاء DataFrame
df = pd.DataFrame(records)

# ترتيب حسب المبيعات (الأكثر مبيعاً أولاً) ثم السعر
df = df.sort_values(by=["المبيعات", "السعر النهائي"], ascending=[False, True]).reset_index(drop=True)

print("[INFO] جاري إنشاء ملف Excel...")

# حفظ إلى Excel أساسي
df.to_excel(OUTPUT_FILE, sheet_name="المنتجات", index=False)

# ============================================================
# STYLING & FORMATTING
# ============================================================

print("[INFO] جاري تطبيق التنسيق والألوان...")

wb = load_workbook(OUTPUT_FILE)
ws = wb.active
ws.title = "المنتجات"

# ── الحدود ──
thin_border = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC')
)

# ── تنسيق الرؤوس ──
header_fill = PatternFill(start_color=HEADER_COLOR, end_color=HEADER_COLOR, fill_type="solid")
header_font = Font(name="Arial", size=12, bold=True, color=HEADER_TEXT)
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_alignment
    cell.border = thin_border

# زيادة ارتفاع رأس الجدول
ws.row_dimensions[1].height = 30

# ضبط عرض الأعمدة وتطبيق التنسيق
column_widths = {
    "A": 8,      # ID
    "B": 35,     # اسم المنتج
    "C": 12,     # السعر الأصلي
    "D": 12,     # السعر النهائي
    "E": 15,     # التوفير
    "F": 15,     # SKU
    "G": 10,     # المبيعات
    "H": 10,     # عدد الصور
    "I": 25,     # الصورة الأولى
    "J": 35,     # الوصف المختصر
    "K": 20,     # الرابط
}

for col_letter, width in column_widths.items():
    ws.column_dimensions[col_letter].width = width

# ── تطبيق الألوان والتنسيق على البيانات ──
for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
    # اختر لون الخط المتبادل
    stripe_color = STRIPE_COLOR_1 if (row_idx - 2) % 2 == 0 else STRIPE_COLOR_2
    
    for col_idx, cell in enumerate(row, start=1):
        cell.fill = PatternFill(start_color=stripe_color, end_color=stripe_color, fill_type="solid")
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
        
        # ── تنسيق محدد لكل عمود ──
        col_letter = get_column_letter(col_idx)
        
        # السعر الأصلي والنهائي — بصيغة عملة
        if col_letter in ["C", "D"]:
            cell.number_format = '0.00'
            if col_letter == "C" and cell.value:  # السعر الأصلي — يشطب
                cell.font = Font(name="Arial", size=10, color="999999", strike=True)
            elif col_letter == "D" and cell.value:  # السعر النهائي — أخضر غامق وجريء
                cell.font = Font(name="Arial", size=11, bold=True, color="27AE60")
        
        # التوفير — أحمر وجريء
        elif col_letter == "E":
            if cell.value:
                cell.font = Font(name="Arial", size=10, bold=True, color="E74C3C")
        
        # SKU
        elif col_letter == "F":
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # المبيعات — بجعله بارز إذا كان عالي
        elif col_letter == "G":
            if cell.value and cell.value > 5:
                cell.font = Font(name="Arial", size=11, bold=True, color="E74C3C")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # عدد الصور
        elif col_letter == "H":
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if cell.value and cell.value > 0:
                cell.font = Font(name="Arial", size=11, bold=True, color=ACCENT_COLOR)
        
        # الرابط — جعله hyperlink أزرق
        elif col_letter == "J":
            if cell.value and str(cell.value).startswith("http"):
                cell.font = Font(name="Arial", size=10, color="0563C1", underline="single")
                original_url = str(row[col_idx-1])
                cell.value = "🔗 اضغط هنا"
                # إضافة hyperlink بشكل صحيح
                ws[cell.coordinate].hyperlink = original_url
        
        # الصورة الأولى — جعل الرابط أزرق
        elif col_letter == "I":
            if cell.value and str(cell.value).startswith("http"):
                cell.font = Font(name="Arial", size=10, color="0563C1", underline="single")
                original_url = str(cell.value)
                cell.value = "📷 عرض"
                # إضافة hyperlink بشكل صحيح
                ws[cell.coordinate].hyperlink = original_url

# ── تجميد الصف الأول (الرؤوس) ──
ws.freeze_panes = "A2"

# ============================================================
# SAVE & SUMMARY
# ============================================================

wb.save(OUTPUT_FILE)

print("\n" + "=" * 60)
print(f"✅ تم الانتهاء بنجاح!")
print(f"📊 الملف: {OUTPUT_FILE}")
print(f"📈 عدد المنتجات: {len(df)}")
try:
    avg_price = pd.to_numeric(df['السعر النهائي'], errors='coerce').mean()
    if not pd.isna(avg_price):
        print(f"💰 متوسط السعر النهائي: {avg_price:.2f}")
except:
    print(f"💰 متوسط السعر النهائي: غير محدد (بيانات قديمة)")
print(f"🏆 أعلى مبيعات: {df['المبيعات'].max()}")
print(f"📸 متوسط الصور: {df['عدد الصور'].mean():.1f}")
print("=" * 60)
print("\n✨ المميزات:")
print("   ✅ رؤوس ملونة احترافية")
print("   ✅ ألوان متبادلة على الصفوف")
print("   ✅ عرض السعر الأصلي والنهائي معاً")
print("   ✅ عرض نسبة الخصم والتوفير")
print("   ✅ روابط قابلة للنقر")
print("   ✅ ترتيب حسب المبيعات")
print("   ✅ صفوف معلقة للرؤوس")
print("\n🎨 الألوان المستخدمة:")
print("   🔵 أزرق غامق: رؤوس الأعمدة")
print("   🟦 أزرق فاتح: خطوط الصفوف")
print("   🟢 أخضر: الأسعار النهائية")
print("   🔴 أحمر: نسب الخصم والمبيعات العالية")
print("   ⚫ رمادي مشطوب: الأسعار الأصلية")
print("\n📝 ملاحظة: قم بتشغيل scraper_details.py لسحب البيانات الجديدة")
print("   لعرض الأسعار والخصومات بشكل صحيح!")
